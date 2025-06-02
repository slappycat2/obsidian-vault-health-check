import sys
import time
from subprocess import Popen

import openpyxl

from openpyxl.styles import Alignment, Font, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from v_chk_cfg_data import *
from v_chk_xl_tabs import *
from v_chk_class_lib import *


# This is defined outside of the Class, to make it easier to use in other scripts.
# Need to restructure this. Should this be a "library", or maybe part of a WB_CLASS?
def xl_clean_cell(cfg, cell_value):
    """
    Replace illegal XML characters in the given cell value with '@'
    if the cell value is a string. Handle datetime objects correctly.
    """
    # Check and process datetime objects directly
    if isinstance(cell_value, datetime):
        cell_value = cell_value.replace(tzinfo=None)

    # Process strings for illegal characters
    if isinstance(cell_value, str):
        # Strip TZ from dates in string format if applicable
        if re.compile(cfg.rgx_noTZdatePattern).search(cell_value):
            cell_value = re.sub(cfg.rgx_noTZdatePattern, cfg.rgx_noTZdateReplace, cell_value)
        return ILLEGAL_CHARACTERS_RE.sub("z", cell_value)

    return cell_value


def xl_a_col(col_num):
    col_alpha = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        col_alpha = chr(65 + remainder) + col_alpha
    return col_alpha

def xl_set_border(ws, cell_range=None, border_type="thin", color="000000"):
    border = Side(border_style=border_type, color=color)

    for row in ws[cell_range]:
        border_first_col = True
        for cell in row:
            cell.border = Border(top=border, bottom=border)
            if border_first_col:
                cell.border = Border(top=border, bottom=border, left=border)
                border_first_col = False
            if row[-1] == cell:
                cell.border = Border(top=border, bottom=border, right=border)

    return

class ExcelExporter:
    def __init__(self, cfg):
        self.tab_def = {}
        self.wb_tabs = {}
        self.Colors = Colors()
        self.c_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.TAG_TOT_COLS = 20  # Set this to the maximum columns to be added to table (w/o the step)
        self.TAG_BEG_COL = 7  # set the to the first column, ie, where to start
        self.PROP_TOT_COLS = 13  # Set this to the maximum columns to be added to table (w/o the step)
        self.PROP_BEG_COL = 9  # set the to the first column, ie, where to start
        self.COL_STEP = 2  # This should be the number of cols that will be printed in each loop
        self.TABLE_GROUPINGS = False
        self.OPEN_ON_CREATE = True

        self.DBUG_AREA51 = True
        self.DBUG_TAB = 'summ'  # DBUG_LVL must be greater than 2
        self.DBUG_LVL = 4
        # self.DBUG_LVL = 0  # Do Not print anything
        # self.DBUG_LVL = 1  # print report level actions only (export, load, save, etc.) + all lower levels
        # self.DBUG_LVL = 2  # print object level actions + all lower levels
        # self.DBUG_LVL = 3  # print export_tab + all lower levels
        # self.DBUG_LVL = 4  # print hdr records + all lower levels
        # self.DBUG_LVL = 5  # print detail records + all lower levels
        # self.DBUG_LVL = 9  # print everything (includes export_cell!)

        self.tab_seq = ['summ', 'pros', 'tags', 'dups']  # ['summ', 'pros', 'tags', 'dups']
        self.tabs_built = {}
        self.wb_tabs_done = {}
        self.ctrl_tots = {
             'pros': 0
            ,'pvals': 0
            ,'tags': 0
            ,'links': 0
            ,'dups': 0
            ,'tag_files': 0
            ,'prop_files': 0
            ,'link_files': 0
            ,'all_files': 0
        }

        cfg = Config()
        cfg = cfg.read_config(cfg)

        if self.DBUG_LVL > 8:
            print(f"ExcelExport - cfg.wb_exec_path: {cfg.wb_exec_path}")

        self.exl_file = Path(cfg.wb_exec_path)

    def initialize_all_tabs(self, wb, tab_list):
        for tab_id in self.tab_seq:
            tab_def = TabDefinition(tab_id)

            tab_name = tab_def['tab_name']
            tab_color = tab_def['tab_color']
            showGridLines = tab_def['showGridLines']
            # if self.DBUG_LVL > 1 and self.DBUG_TAB == self.tab_def[tab_id]:
            #     print(f"Initializing Tab: {tab_id}")  # Set a breakpount on this line

            # Create the tab, or rename Sheet 1, in the Summaries case...
            if tab_id == 'summ':
                tab = wb.active
                tab.title = tab_name
            else:
                tab = wb.create_sheet(title=tab_name)  # Excel allows max 31 characters in tab names

            self.wb_tabs[tab_id] = tab

            tab.sheet_properties.tabColor = tab_color
            tab.sheet_view.showGridLines = showGridLines

        return wb

    # =================================================================================
    def export(self, cfg):
    # =================================================================================
        if self.DBUG_LVL > 1:
            print(f"ExcelExport.export - cfg.wb_exec_path: {cfg.wb_exec_path}")

        # Create the workbook instance
        wb = openpyxl.Workbook()

        wb = self.initialize_all_tabs(wb, self.tab_seq)

        # Initialize ALL tabs (not Area51) because we may need to update more than one at a time
        for tab_id in self.tab_seq:
            if self.DBUG_LVL > 3:
                print(f"Exporting tab: {tab_id}")

            self.tab_def = TabDefinition(tab_id)
            # self.tab_def = self.wb_tabs[tab_id]
            self.tabs_built[tab_id] = self.tab_def
            self.wb_tabs_done[tab_id] = self.export_tab(cfg, wb, self.tab_def)

        # self.export_summary(cfg, wb, self.wb_tabs_done['summ'], self.tabs_built['summ'])
        if self.DBUG_AREA51:
            self.export_area51(cfg, wb)

        self.save_workbook(cfg, wb)

    def save_workbook(self, cfg, wb):
        if self.DBUG_LVL > 8:
            print(f"Saving Spreadsheet: {cfg.v_chk_xls_pname}")

        # save and load workbook
        if os.path.isfile(cfg.v_chk_xls_pname):
            user_fixed = 'n'
            w_time = 2 # secs
            retry_max = 2
            retry_count = 0
            while user_fixed.lower() in ["y", "n"]:
                try:
                    os.remove(cfg.v_chk_xls_pname)
                    user_fixed = ''
                except PermissionError:
                    print(f"File {cfg.v_chk_xls_pname} must be closed. Trying again in {w_time} second...")
                    time.sleep(w_time)
                    if user_fixed.lower() not in ["y", "n"]:
                        if retry_count < retry_max:
                            user_fixed = 'y'
                            retry_count += 1

                        else:
                            print(f"File {cfg.v_chk_xls_pname} must be closed. Max Retries {retry_max} exceed. Exiting...")
                            sys.exit(1)

                except Exception as e:
                    print(f"Error removing file {cfg.v_chk_xls_pname}: {e}")
                    user_fixed = input(f"Try again? (Y/n): ")
                    if user_fixed.lower() not in ["y", "n"]:
                        user_fixed = 'y'

        wb.save(cfg.v_chk_xls_pname)

        if self.OPEN_ON_CREATE:
            self.load_spreadsheet(cfg)

    def load_spreadsheet(self, cfg):
        pid = Popen([cfg.wb_exec_path, cfg.v_chk_xls_pname]).pid
        return pid

    def export_cell(self, cfg, tab, col_def_list, val, row_idx):
        # col,row,font,sz, w,t_clr,f_clr,Bold,Ital,  Align,  val ] = 11
        col_idx, def_row, c_font, c_sz, col_w, txt_clr, fill_clr, bold_bool, ital_bool, align_val, def_val = col_def_list

        if self.DBUG_LVL > 8 and self.DBUG_TAB == self.tab_def['tab_id']:
            print(f"Exporting Cell (Col,Row): {col_def_list[0]},{row_idx} val:{val}  def_val: {def_val}") #  Set a breakpount on this line

        if self.DBUG_LVL > 8 and self.DBUG_TAB == self.tab_def['tab_id']:
            print(f"Exporting Cell (Col,Row): {col_def_list[0]},{row_idx} val:{val}  def_val: {def_val}") #  Set a breakpount on this line

        hyperlink = False
        if isinstance(val, str) and val.upper().startswith("=HYPERLINK"):
            hyperlink = True

        if isinstance(val, str) and val.endswith(":"):
            txt_clr = self.Colors.clr_red20
            ital_bool = True
            bold_bool = True

        if val is None or val == "":
            val = def_val

        if row_idx == 0:
            row_idx = def_row

        if c_sz is None or c_sz == 0:
            c_sz = 10

        if c_font == '' or c_font is None:
            c_font = "default"

        fg_clr = txt_clr
        if fill_clr != "" and fill_clr is not None:
            fg_clr = self.Colors.Code_LOV[fill_clr][1]

        val = xl_clean_cell(cfg, val)

        cell = tab.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name=c_font, size=c_sz, bold=bold_bool, italic=ital_bool)
        cell.alignment = Alignment(horizontal=align_val, vertical="center")
        if self.DBUG_LVL > 8 and row_idx < 13:
            print(f"Set Cell: at {row_idx}:{col_idx} to:{val}")

        if txt_clr != "" and txt_clr is not None:
            cell.font = Font(name=c_font, size=c_sz, color=txt_clr, bold=bold_bool, italic=ital_bool)

        if hyperlink:
            if txt_clr == "" or txt_clr is None:
                txt_clr = self.tab_def['tab_link_clr']
            cell.font = Font(color=txt_clr, underline='single', bold=True, italic=ital_bool)

        if fill_clr != "":
            cell.fill = PatternFill(start_color=fill_clr, end_color=fill_clr, fill_type="solid", fgColor=fg_clr)

        # col width
        a_col = xl_a_col(col_idx)
        if col_w > 0:
            tab.column_dimensions[a_col].width = col_w

        return col_idx, cell

    #   export_tab(wb, cfg, 'pros')
    def export_tab(self, cfg, wb, tab_def):
        tab_id = tab_def['tab_id']
        tab = self.wb_tabs[tab_id]

        tab_comments = tab_def['tab_comments']
        tab_notes = tab_def['tab_notes']
        tab_table_style = tab_def['tab_table_style']
        tab_color = tab_def['tab_color']
        tbl_hdr_row = tab_def['tbl_hdr_row']
        tbl_beg_col = tab_def['tbl_beg_col']
        tbl_end_col = tab_def['tbl_end_col']
        tab_table_links_cols = tab_def['tab_table_links_cols']
        tab_tots_isVisible_col = tab_def['tab_tots_isVisible_col']
        tbl_name = tab_def['tbl_name']
        tab_cd_table_hdr = tab_def['tab_cd_table_hdr']
        tab_cd_table_dtl = tab_def['tab_cd_table_dtl']
        hdr_IsVis        = tab_def['hdr_IsVis']
        showGridLines    = self.tab_def['showGridLines']
        if self.DBUG_LVL > 2 and self.DBUG_TAB == self.tab_def['tab_id']:
            print(f"Exporting Tab: {tab_id}") #  Set a breakpount on this line

        # Create the tab, or rename Sheet 1, in the Summaries case...
        # if tab_id == 'summ':
        #     tab = wb.active
        #     tab.title = tab_name
        # else:
        #     tab = wb.create_sheet(title=tab_name)  # Excel allows max 31 characters in tab names

        # tab.sheet_properties.tabColor = tab_color
        # tab.sheet_view.showGridLines = showGridLines

        # TAB HEADINGS export the Tabs Main Title in a large font
        temp_def = self.tab_def['tab_title_def']
        row_idx = temp_def[1]
        _, cell = self.export_cell(cfg, tab, temp_def, '', row_idx)

        # notes
        if len(tab_notes) > 0:
            temp_def = self.tab_def['tab_notes_def']
            row_idx = temp_def[1]
            for row_idx, comment in enumerate(tab_notes, start=temp_def[1]):
                _, cell = self.export_cell(cfg, tab, temp_def, comment, row_idx)

        # comments
        if len(tab_comments) > 0:
            temp_def = self.tab_def['tab_comments_def']
            for row_idx, comment in enumerate(tab_comments, start=temp_def[1]):
                _, cell = self.export_cell(cfg, tab, temp_def, comment, row_idx)

        # TABLE-Hdr export the main body table - column headers
        col_idx = tbl_beg_col
        row_idx = tbl_hdr_row
        for hdr_key, hdr_col_def_list in tab_cd_table_hdr.items():

            # unpack header definitions
            col_idx, cell = self.export_cell(cfg, tab, hdr_col_def_list, '', tbl_hdr_row)
            if self.DBUG_LVL == 8:
                print(f"Processing [Header]:{hdr_key}")
                print(f" {tab_id}  row:{tbl_hdr_row}  col: {col_idx}  list: {hdr_col_def_list}")

        tab[f"{xl_a_col(tbl_beg_col + 1)}{tbl_hdr_row + 1}"] = "Nothing Found."
        # tab[f"E{tbl_hdr_row + 1}"].font = tab[f"E{tbl_hdr_row}"].font = Font(bold=True)

        # TABLE-Dtl Begin Detail
        row_idx = tbl_hdr_row + 1
        p_v_Index_p_count, p_v_Index_v_count = 0, 0
        beg_prop_group = row_idx + 1
        last_prop_name = ""
        pros_total = values_total = files_total = 0
        # TABLE-Dtl Loop each prop_name, getting the list of all values
        for prop_name, values_dict in sorted(cfg.pros.items()):
            if tab_id == 'pros' and (prop_name == "tags" or prop_name == "tags:" or  prop_name == 'xkey_dup_files'):
                continue
            if tab_id == 'tags' and prop_name != "tags" and prop_name != "tags:":
                continue
            if tab_id == 'dups' and prop_name != 'xkey_dup_files':
                continue
            if tab_id == 'summ' and prop_name == 'xkey_dup_files':
                continue

            # so for dups, values_dict will be all pathnames using this filename
            prop_name = xl_clean_cell(cfg, prop_name)
            p_v_Index_p_count += 1
            pros_total += 1
            p_v_Index_v_count = 0
            if tab_id == "summ" and last_prop_name == "":
                last_prop_name = prop_name

            sorted_values_dict = values_dict.items()
            # TABLE-Dtl Now, build vals[], by looping each value,  getting a list of Files-Where-Used
            # for value_item, value_files_list in values_dict.items():
            for value_item, value_files_list in sorted_values_dict:
                if tab_id == 'dups' and len(value_files_list) == 1:
                    continue

                # First, define what the values are going to be for this row
                value_item_count = len(value_files_list)
                vals = []
                p_v_Index_v_count += 1
                values_total += 1
                files_total += value_item_count

                # TABLE Dtl - Set up First (Fixed) Columns Values
                if tab_id == "pros":
                    vals = [int((row_idx - tbl_hdr_row))
                        , prop_name
                        , value_item
                        , value_item_count
                        , f"{p_v_Index_p_count:03d}-{p_v_Index_v_count:05d}"
                            ]
                if tab_id == "tags":
                    vals = [int((row_idx - tbl_hdr_row))
                        , value_item
                        , value_item_count
                            ]
                if tab_id == "summ":
                    vals = [int((row_idx - tbl_hdr_row))
                        , last_prop_name
                        , values_total
                        , files_total
                            ]
                if tab_id == "dups":
                    vals = [int((row_idx - tbl_hdr_row))
                        , value_item
                        , value_item_count
                            ]
                # TABLE-Dtl Set up List of Files Used, convert them to obsidian URLs, store in vals[]
                this_row_dtl = tab_cd_table_dtl
                # tab_def["tab_table_links"] =  [9, 0, '', 11, 18, "", "", False, False, 'left'  , '']
                for col_count, file in enumerate(value_files_list, start=1):
                    if col_count > tab_table_links_cols:
                        break

                    if tab_id == "dups":
                        obs_file = str(file)
                        file_link = f'=hyperlink("obsidian://open?vault=o2&file={obs_file}","{obs_file.replace(".md", "")}")'
                    else:
                        obs_file = Path(file).name
                        file_link = f'=hyperlink("obsidian://open?vault=o2&file={obs_file}","{obs_file}")'

                    vals = vals + [file_link, " "]


                # TABLE-Dtl Finally, now we have all the col values in vals, so we can add this row...
                if self.DBUG_LVL > 4:
                    print(f"Processing [Property]:{prop_name} - [Value]:{value_item} - [Count]: {value_item_count}")

                # TABLE-Dtl Start of a new row print, loop thru tab_table columns and add to sheet
                if tab_id == "summ" and prop_name == last_prop_name:
                    continue

                for dbug_only, dtl_col_def_list in this_row_dtl.items():
                    col_idx = dtl_col_def_list[0]

                    # TABLE-Dtl If there is no value in vals for this col
                    if len(vals) <= (col_idx - tbl_beg_col):
                        continue

                    val = vals[col_idx - tbl_beg_col]

                    if self.DBUG_LVL > 8:
                        print(f"Col: {col_idx}-{dbug_only}\t\t  "
                              f"Val {col_idx - tbl_beg_col}: {vals[col_idx - tbl_beg_col]}")

                    if col_idx == tab_tots_isVisible_col:
                        val = tab_def['f_isVisible']

                    col_idx, cell = self.export_cell(cfg, tab, dtl_col_def_list, val, row_idx)

                    # if col_idx == 5 and prop_name == last_prop_name and prop_name == "pros":
                    #     cell.font = Font(color=self.tab_def.clr_gry35)

                if self.TABLE_GROUPINGS:
                    if prop_name != last_prop_name and last_prop_name != "":
                        if beg_prop_group != row_idx - 1:
                            tab.row_dimensions.group(beg_prop_group, row_idx - 1, hidden=False)
                            # wgh = openpyxl.worksheet.dimensions.DimensionHolder(tab, reference='index', default_factory=None)
                            # wgh.group(beg_prop_group, row_idx - 1, outline_level=1, hidden=False)
                            beg_prop_group = row_idx + 1

                last_prop_name = prop_name
                values_total = 0
                files_total = 0
                row_idx += 1

        # Group last group
        if self.TABLE_GROUPINGS:
            if beg_prop_group != row_idx - 1:
                tab.row_dimensions.group(beg_prop_group, row_idx - 1, hidden=False)

        # tab.row_dimensions.co(outlineLevel=1, worksheet=tab, collapsed=True)
        # create a table
        # Calculate table size
        # last_row = tab.max_row
        # if last_row != tbl_hdr_row + 1:
        #     print(f"last_row: {last_row} <> row_idx - tbl_hdr_row: {row_idx - tbl_hdr_row}")
        #     _ = input(f"Press  Return to continue...")
        ret = self.format_as_table(tab, tbl_name, tab_table_style, row_idx)
        last_row = row_idx

        # Add formulas for totals - tab_cd_fixed_grid
        # 'tab_cd_fixed_grid': {  # col, row, sz, t_clr, f_clr, Bold, Italic, Align, Value
        #     'Analysis': [10, 2, 14, "", "", True, False, 'center', '']
        #     , 'Property': [12, 2, 12, "", "", True, False, 'center', '']
        #     , 'Values': [14, 2, 12, "", "", True, False, 'center', '']
        #     , 'Files Used': [16, 2, 12, "", "", True, False, 'center', '']
        #     , 'Unique Values': [10, 3, 12, "", "", True, False, 'left', '']
        #     , 'Column Totals': [10, 4, 12, "", "", True, False, 'left', '']
        #     , 'x-null-values': [10, 5, 12, red, "", True, True, 'left',
        #                         '=IFERROR(IF(AGGREGATE(3,3,tbl_pros[Values])<>SUM(tbl_pros[IsVisible]),"Empty Values Detected!",""),"")']
        #     , 'x-filters-on': [10, 6, 12, red, "", True, True, 'left',
        #                        '=IFERROR(IF(COUNTA(tbl_pros[P-V Index])<>SUM(tbl_pros[IsVisible]),"Column filters applied--Totals now reflect column filters!",""),"")']
        #     , 'x-uniq-prop': [12, 3, 11, "", "", False, False, 'center',
        #                       '=COUNTA(UNIQUE(FILTER(tbl_pros[Property],tbl_pros[IsVisible])))']
        #     , 'x-uniq-val': [14, 3, 11, "", "", False, False, 'center',
        #                      '=COUNTA(UNIQUE(FILTER(tbl_pros[Values],tbl_pros[IsVisible])))']
        #     , 'x-ctot-prop': [12, 4, 11, "", "", False, False, 'center', '=AGGREGATE(3,3,tbl_pros[Property])']
        #     , 'x-ctot-val': [14, 4, 11, "", "", False, False, 'center', '=AGGREGATE(3,3,tbl_pros[Row])']
        #     , 'x-ctot-files': [16, 4, 11, "", "", False, False, 'center', '=AGGREGATE(9,3,tbl_pros[Times Used])']

        # Export fixed grid cells
        try:

            tots_grid_key = 'tab_cd_fixed_grid'  # this is also used elsewhere ;)
            tot_table = tab_def[tots_grid_key]
            vis_key = 'isVisible'
            tab_table_link_spcrs = self.tab_def['tab_table_link_spcrs']
            if tab_tots_isVisible_col:
                # STEP 1 - Set IsVisible formula in last table column
                # Set isVisible Column Header

                row_num = tbl_hdr_row
                cell_def = tab_def[tots_grid_key][vis_key]

                col, row, font, sz, w, t_clr, fill_clr, Bold, Ital, Align, val = cell_def
                _, cell = self.export_cell(cfg, tab, cell_def, hdr_IsVis, row_num)
                row_num += 1

                # Now, export IsVisible Formula to all cells in the IsVisible column
                for row_num in range(row_num, last_row):
                    _, cell = self.export_cell(cfg, tab, cell_def, val, row_num)

            # Now, print the rest of the grid

            val = None    # get rid of the last val
            row_idx = 0     # use the values defined in the grid
            col_idx = 0
            # ========================================================================
            # export Totals Grid, both headers and formulas for totals
            # ========================================================================

            for tot_key, tot_col_def_list in tot_table.items():
                if tot_key == vis_key:
                    continue

                if self.DBUG_LVL > 3: # and self.DBUG_TAB == self.tab_def['tab_id']:
                    print(f"Grid: '{tot_key}' {tot_col_def_list}")
                    # ,{row_idx} val:{val}  def_val: {def_val}")  # Set a breakpount on this line

                # def export_cell(self, cfg, tab, col_def_list, val, row_idx):
                #     col_idx, col_w, txt_clr, fill_clr, bold_bool, align_val = col_def_list
                #   col, row, sz, t_clr, fill_clr, Bold, Italic, Align, Value          len=9
                #   col, w, t_clr, fill_clr, Bold, Align                               len=6
                _, cell = self.export_cell(cfg, tab, tot_col_def_list, val, row_idx)

                val = None

        except KeyError:
            pass

        except Exception as e:
            print(f"ERROR: {e}")
            print(f"ERROR: An Unhandled error occurred in tots_grid table processing.")
            _ = input(f"Press  Return to continue...")

        return tab

    def format_as_table(self, tab, tbl_nm, tab_tbl_style, tot_rows):
        tbl_beg_col   = self.tab_def['tbl_beg_col']
        tbl_end_col     = self.tab_def['tbl_end_col']
        tbl_hdr_row = self.tab_def['tbl_hdr_row']
        tbl_rng = f"{xl_a_col(tbl_beg_col)}{tbl_hdr_row}:{xl_a_col(tbl_end_col)}"
        if tot_rows == int((tot_rows - tbl_hdr_row)):
            tbl_rng = f"{tbl_rng}11"
        else:
            tbl_rng = tbl_rng + str(tot_rows - 1)

        if self.DBUG_LVL > 5:
            print(f"tbl_name: {tbl_nm}  tbl_rng: {tbl_rng}")

        tbl = Table(displayName=tbl_nm, ref=tbl_rng)
        tbl_style = TableStyleInfo(name=tab_tbl_style, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = tbl_style
        tab.add_table(tbl)

        return tab

    def export_dups(self, cfg, wb):
        # self.prop_tab_dict{prop_name: [prop_vals_list]}
        dup_tab = wb.create_sheet(title="Dups")  # Excel allows max 31 characters in tab names

        dup_tab.sheet_properties.tabColor = cfg.clr_ora64
        dup_tab.sheet_properties.tabColor = self.Colors.clr_ora64

        cell = dup_tab.cell(row=2, column=5, value=f'Total Duplicates')
        cell.font = Font(size=12, italic=True, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        hdr = "Duplicate Files (Identical md-filenames listed under different directories)"

        dup_tab["D4"] = hdr
        dup_tab["D4"].font = Font(size=14, bold=True)

        # Write header at D10
        tbl_hdr_row = 10
        row_idx = 1
        dup_tab[f"D{tbl_hdr_row}"] = "RowId"
        dup_tab[f"E{tbl_hdr_row}"] = "Filename"
        dup_tab.column_dimensions['E'].width = 45
        dup_tab.column_dimensions['F'].width = 80
        dup_tab.column_dimensions['G'].width = 80

        # assume no files and make a note, which will be overwritten if 1 exists
        dup_tab[f"E{tbl_hdr_row + 1}"] = "No Duplicates"
        dup_tab[f"E{tbl_hdr_row + 1}"].font = dup_tab[f"E{tbl_hdr_row}"].font = Font(bold=True)

        # Write content under header
        highest_dup_count = 0
        for file, dir_list in sorted(cfg.dup_files.items()):
            if len(dir_list) == 1:
                continue
            filename = xl_clean_cell(cfg, file)

            # D - Row
            cell = dup_tab.cell(row=row_idx + tbl_hdr_row, column=4, value=row_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # E - Property Name
            cell = dup_tab.cell(row=row_idx + tbl_hdr_row, column=5, value=filename)
            cell.alignment = Alignment(horizontal="left", vertical="center")

            dup_count = 0
            for filepath in sorted(dir_list):
                dup_count += 1
                if dup_count > 11:
                    continue
                if dup_count > 10:
                    filename = "...plus more"
                if self.DBUG_LVL > 2:
                    print(f"Create excel - processing duplicates filename: {filename}")
                # F - Values Count
                filepath = xl_clean_cell(cfg, filepath)
                cell = dup_tab.cell(row=row_idx + tbl_hdr_row, column=dup_count + 5, value=filepath)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if dup_count > highest_dup_count:
                    highest_dup_count = dup_count
                    cell = dup_tab.cell(row=tbl_hdr_row, column=dup_count + 5, value=f"Dup {dup_count}")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            row_idx += 1

        cell = dup_tab.cell(row=2, column=4, value=row_idx)
        cell.font = Font(size=14, bold=True)

        last_col = dup_tab.max_column

        # create a table
        # Calculate table size
        tbl_name = "tbl_dups"
        last_col = chr(last_col + 64)
        if row_idx == 0:
            tbl_rng = "D10:G11"
        else:
            tbl_rng = "D10:" + last_col + str(row_idx + 9)
        if cfg.dbug:
            print(f"tbl_name: {tbl_name}  tbl_rng: {tbl_rng}")

        tab = Table(displayName=tbl_name, ref=tbl_rng)
        style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        dup_tab.add_table(tab)

        return

    def export_area51(self, cfg, wb):
        # Create Notes tab with color index
        # Step 5: Create "Area51" tab
        area51_tab = wb.create_sheet(title="Area51")
        area51_tab.sheet_properties.tabColor = 'FF0000'
        area51_tab.font = Font(color='FFFFFF')
        area52_tab = wb.create_sheet(title="Area52")
        area52_tab.sheet_properties.tabColor = 'FF0000'
        area52_tab.font = Font(color='FFFFFF')
        colors_tab = wb.create_sheet(title="Colors")
        colors_tab.sheet_properties.tabColor = 'FF0000'
        colors_tab.font = Font(color='FFFFFF')

        # Dump tabs_built
        tab_col_idx = 2
        tab_col_step = 6
        hdr_row = 15
        a51_def = {}

        a52_tab_col_idx = 2
        a52_tab_col_step = 14
        a52_hdr_row = 10
        a52_sz = 12
        a52_def = {}
        a52_def = self.a52_reset_defs(a52_def, hdr_row, a52_sz)
        # start a new tab_def set of cols
        compare_tab = 'pros'
        sub_vals = {'tab_cd_table_hdr': 'hdr'
                  , 'tab_cd_table_dtl': 'dtl'
                  , 'tab_cd_fixed_grid': 'grid'}
        col_def_names = ['col', 'row', 'font', 'sz', 'w', 't_clr', 'fill_clr',
                         'Bold', 'Ital', 'Align','val']
        col_defs = {}
        c_tab = self.tabs_built[compare_tab].tab_def
        # A51: Start a new table for a new tab
        for tab_id in self.tab_seq:
            # A51: start a new tab_def set of cols
            col_idx = tab_col_idx
            start_col = tab_col_idx
            sz = 12
            row_idx = hdr_row
            tbl_rng = f"{xl_a_col(col_idx)}{row_idx}:{xl_a_col(col_idx + 3)}"
            a51_def = self.a51_reset_defs(a51_def, hdr_row, sz)
            # A51: Print next tab set header (loops 4 cols)
            for col, cell_def in a51_def['tab_cd_table_hdr'].items():
                cell_def[0] = col_idx
                _, _ = self.export_cell(cfg, area51_tab, cell_def, '', row_idx)

                col_idx += 1

            # A51: start of new tab set details
            col_idx = start_col
            row_idx = hdr_row
            row_id = 1

            a51_tab_def = self.tabs_built[tab_id].tab_def
            sub_defs = {}
            a52_sub_defs = {}

            # A51: export Main tab_def dtl
            for key, value in a51_tab_def.items():
                # a51_item = a51_tab_def
                col_idx = start_col
                col_idx, row_idx, row_id = (
                    self.a51_export_row(
                              cfg
                            , area51_tab
                            , a51_def
                            , key
                            , value
                            , col_idx
                            , row_idx
                            , row_id
                            , hdr_row
                            , sz
                    ))

                if key in sub_vals:
                    # sd_key = f"{row_id}-{sub_vals[key]}"
                    sd_key = f"{row_id}-{sub_vals[key]}"
                    a52key = f"{sub_vals[key]}"
                    sub_defs[sd_key] = value
                    a52_sub_defs[a52key] = value


            # A51: loop sub_vals
            for xkey, xval in sub_defs.items():
                for skey, sub_val in xval.items():

                    # a51_def = self.a51_reset_defs(a51_def, hdr_row, sz)

                    skey = f"  {xkey}: {skey}"
                    col_idx = start_col
                    col_idx, row_idx, row_id = (
                            self.a51_export_row(
                                cfg
                                , area51_tab
                                , a51_def
                                , skey
                                , sub_val
                                , col_idx
                                , row_idx
                                , row_id
                                , hdr_row
                                , sz
                            )
                        )
            # A51: Create A51 Table
            tbl_nm = f"tbl_A51_{tab_id}"
            tab_tbl_style = self.tabs_built[tab_id].tab_def['tab_table_style']
            tbl_rng = f"{tbl_rng}{row_idx}"
            tbl = Table(displayName=tbl_nm, ref=tbl_rng)
            tbl_style = TableStyleInfo(name=tab_tbl_style, showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = tbl_style
            area51_tab.add_table(tbl)

            # ============================================
            # Area52: Print Cell Defs
            # ============================================

            # Next, print the cell_defs (cd)...
            # Area52: Dump cell_defs

            a52_def = self.a52_reset_defs(a52_def, a52_hdr_row, a52_sz)
            # Print next tab set header (loops 4 cols)
            col_idx = a52_tab_col_idx
            row_idx = a52_hdr_row
            tbl_rng = f"{xl_a_col(col_idx)}{row_idx}:{xl_a_col(col_idx + 11)}"

            if self.DBUG_LVL > 1:
                print(f"\n\nA52 {tab_id} Header --------------------------------------------")
            for col, cell_def in a52_def['tab_cd_table_hdr'].items():
                if self.DBUG_LVL > 7:
                    print(f"A52 Sub-Next Tab-Set Heading print=col:  {col} cell_def: {cell_def}")
                cell_def[0] = col_idx
                if col_idx == a52_tab_col_idx:
                    cell_def[10] = tab_id
                _, _ = self.export_cell(cfg, area52_tab, cell_def, '', row_idx)
                col_idx += 1

            # row_idx += 1
            # start of new tab set details
            # col_defs[idx = start_col
            # value eg. {'Row': [10, 10, '', 12, 8, '', '', True, False, 'center', 'RowId'],
            #            'Prop': [11, 10, '', 12, 22, '', '', True, False, 'left', 'Property'],
            #            'Values Count': [12, 10, '', 12, 10, '', '', True, False, 'center', 'Values'],
            #            'File Count': [13, 10, '', 12, 10, '', '', True, False, 'center', 'Files']}
            #   col_def_names = ['col', 'row', 'font', 'sz', 'w', 't_clr', 'fill_clr',
            #                     'Bold', 'Ital', 'Align','val']  # len = 11
            #   sub_vals = {'tab_cd_table_hdr': 'hdr'
            #             , 'tab_cd_table_dtl': 'dtl'
            #             , 'tab_cd_fixed_grid': 'grid'}
            # Therefore, sub_defs will be something like:
            #   hdr: {'Row': [10, 10, '', 12, 8, '', '', True, False, 'center', 'RowId'],
            #         'Prop': [11, 10, '', 12, 22, '', '', True, False, 'left', 'Property'],
            #         'Values Count': [12, 10, '', 12, 10, '', '', True, False, 'center', 'Values']}
            # sub_val

            # A52: loop sub_vals
            if self.DBUG_LVL > 1:
                print(f"\nA52 Outer --------------------")

            for xkey, xval in a52_sub_defs.items():
                xkey = f"{tab_id}-{xkey}"

                if self.DBUG_LVL > 1:
                    print(f"Sub-Next Tab-Loop Sub-Vals=xkey:  {xkey} xval: {xval}")

                col_idx = a52_tab_col_idx
                # col_idx, row_idx, row_id = (
                #     self.a52_export_row(
                #           cfg
                #         , area52_tab
                #         , a52_def
                #         , skey
                #         , sub_val
                #         , col_idx
                #         , row_idx
                #         , row_id
                #         , a52_hdr_row
                #         , a52_sz
                #     )
                # )
                if self.DBUG_LVL > 1:
                    print(f"A52 Inner loop ---------")

                for skey, sub_val in xval.items():

                    # a52_def = self.a52_reset_defs(a52_def, hdr_row, sz)

                    skey = f"{xkey}: {skey}"
                    if self.DBUG_LVL > 1:
                        print(f"A52 Sub-Inner Sub-Loop Sub-Vals=skey:  {skey} sub_val: {sub_val}")

                    col_idx = a52_tab_col_idx
                    col_idx, row_idx, row_id = (
                            self.a52_export_row(
                                cfg
                                , area52_tab
                                , a52_def
                                , skey
                                , sub_val
                                , col_idx
                                , row_idx
                                , row_id
                                , a52_hdr_row
                                , a52_sz
                            )
                        )

            # Create A52 Table
            #  tab_tbl_style = self.tabs_built[tab_id].tab_def['tab_table_style']
            tbl_nm = f"tbl_A52_{tab_id}"
            tbl_rng = f"{tbl_rng}{row_idx}"
            tbl = Table(displayName=tbl_nm, ref=tbl_rng)
            tbl_style = TableStyleInfo(name=tab_tbl_style, showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = tbl_style
            area52_tab.add_table(tbl)

            tab_col_idx += tab_col_step
            a52_tab_col_idx += a52_tab_col_step



        # ==========================================
        # all done! Now, just print the color chart
        # ==========================================
        self.a51_dump_colors(cfg, wb, colors_tab)

    def a52_export_row(self
                            , cfg
                            , area52_tab
                            , a52_def
                            , key
                            , value
                            , col_idx
                            , row_idx
                            , row_id
                            , hdr_row
                            , sz
                       ):
        vals = value
        vals.insert(0, key)    # list1.insert(4, 10)
        row_idx += 1
        row_id += 1
        val = 0

        # A52: export_row - get class and length of this cell
        # export_row: get value class
        cls_type = value.__class__.__name__
        if cls_type in ['str', 'list', 'dict', 'tuple', 'set']:
            length = len(value)
        else:
            length = "n/a"

        # value = f'{value}' # to avoid errors on if's, below

        # vals[1] = key
        # vals[2] = f'{cls_type} ({length}))'
        # vals[3] = f'{value}'

        # A52: export_row - loop the 12 cells
        for _, cell_def in a52_def['tab_cd_table_dtl'].items():
            cell_def[0] = col_idx

            # A52: Handle special values
            # A52: color
            if (val == 2 and vals[3] in self.Colors.Code_LOV):     # and (len(c_val) == 8 or len(c_val) == 6)
                cell_def[6], cell_def[5] = self.Colors.Code_LOV[vals[3]]

            # # A52: formulae

            cls_val = vals[val].__class__.__name__
            if cls_val == 'str' and vals[val].startswith("="):
                 vals[val] = f"'{vals[val]}"

            cell_def[10] = vals[val]
            if self.DBUG_LVL > 7:
                print(f"col_idx: {col_idx}  row_idx: {row_idx} cls_type: '{cls_type}' cell_def: {cell_def}")
            _, _ = self.export_cell(cfg, area52_tab, cell_def, '', row_idx)
            val += 1
            col_idx += 1
            cell_def[5] = ''
            cell_def[6] = ''

        return col_idx, row_idx, row_id


    def a51_export_row(self
                            , cfg
                            , area51_tab
                            , a51_def
                            , key
                            , value
                            , col_idx
                            , row_idx
                            , row_id
                            , hdr_row
                            , sz
                       ):

        vals = [None, None, None, None, " "]
        val = 0
        row_idx += 1
        row_id += 1

        # A51: export_row: get value class
        cls_type = value.__class__.__name__
        if cls_type in ['str', 'list', 'dict', 'tuple', 'set']:
            length = f'{len(value)}'
        else:
            length = "n/a"

        value = f'{value}' # A51: to avoid errors on if's, below

        vals[0] = row_id
        vals[1] = key
        vals[2] = f'{cls_type} ({length})'
        vals[3] = f'{value}'

        # A51: export_row: loop the 4 cells
        for _, cell_def in a51_def['tab_cd_table_dtl'].items():
            cell_def[0] = col_idx
            c_val = vals[val]

            # A51: Handle special values
            # A51: color
            if (val == 2 and vals[3] in self.Colors.Code_LOV):     # and (len(c_val) == 8 or len(c_val) == 6)
                clr_list = self.Colors.Code_LOV[vals[3]]
                cell_def[5] = clr_list[1]
                cell_def[6] = clr_list[0]
            if val == 3 and key.startswith("f_"):
                vals[3] = f"'{vals[3]}"

            cell_def[10] = vals[val]

            _, _ = self.export_cell(cfg, area51_tab, cell_def, '', row_idx)
            val += 1
            col_idx += 1
            cell_def[5] = ''
            cell_def[6] = ''

        return col_idx, row_idx, row_id

    def a51_reset_defs(self, a51_def, hdr_row, sz):

        a51_def['tab_cd_table_hdr'] = {
            # A51: col,row,font,sz, w,t_clr,fill_clr,Bold,Ital,  Align,  val ] = 11
              "Row":  [0, hdr_row, '', sz,  4, "", "", True, False, 'center', "RowId"]
            , "Key":  [0, hdr_row, '', sz, 25, "", "", True, False, 'left',   "[Key]"]
            , "Type": [0, hdr_row, '', sz, 10, "", "", True, False, 'center', "Type (Len)"]
            , "Val":  [0, hdr_row, '', sz, 70, "", "", True, False, 'left',   "[Value]"]
            , "Spc":  [0, hdr_row, '',  8,  1, "", "", True, False, 'right',  "Spc   "]
        }
        a51_def['tab_cd_table_dtl'] = {
            # A51: col,row,font,sz, w,t_clr,fill_clr,Bold,Ital,  Align,  val ] = 11
              "Row":  [0, 0,       '', sz, 0, "", "", False, False, 'center', ""]
            , "Key":  [0, 0,       '', sz, 0, "", "", True,  False, 'left', ""]
            , "Type": [0, 0,       '', sz, 0, "", "", False, True,  'center', ""]
            , "Val":  [0, 0,       '', sz, 0, "", "", False, False, 'left', ""]
            , "Spc":  [0, hdr_row, '', sz, 0, "", "", True,  False, 'right', " "]

        }

        return a51_def

    def a52_reset_defs(self, a52_def, hdr_row, sz):
        a52_def['tab_cd_table_hdr'] = {
            # A52: col,row,font,sz, w,t_clr,fill_clr,Bold,Ital,  Align,  val ] = 11
              'skey':  [0, hdr_row, '', sz, 25, "", "", False, False, 'left', ""]
            , 'row':   [0, hdr_row, '', sz,  4, "", "", False, False, 'center', "Row"]
            , 'col':   [0, hdr_row, '', sz,  4, "", "", False, False, 'center', "Col"]
            , 'font':  [0, hdr_row, '', sz, 20, "", "", False, False, 'left',   "Font"]
            , 'sz':    [0, hdr_row, '', sz,  4, "", "", False, False, 'center', "Sz"]
            , 'w':     [0, hdr_row, '', sz,  4, "", "", False, False, 'center', "w"]
            , 't_clr': [0, hdr_row, '', sz, 10, "", "", False, False, 'center', "Text Clr"]
            , 'fill_clr': [0, hdr_row, '', sz, 10, "", "", False, False, 'center', "Fill Clr"]
            , 'Bold':  [0, hdr_row, '', sz,  7, "", "", False, False, 'center', "Bold"]
            , 'Ital':  [0, hdr_row, '', sz,  7, "", "", False, False, 'center', "Italic"]
            , 'Align': [0, hdr_row, '', sz,  7, "", "", False, False, 'center', "Align"]
            , 'val':   [0, hdr_row, '', sz, 20, "", "", False, False, 'left',   "Dflt Val"]
        }
        a52_def['tab_cd_table_dtl'] = {
            # A52:    col,row,font,sz, w,t_clr,fill_clr,Bold, Ital,   Align,  val ] = 11
              'skey':  [0, 0, '', sz, 25, "",   "", False, False, 'left',   ""]
            , 'row':   [0, 0, '', sz,  4, "",   "", False, False, 'center', ""]
            , 'col':   [0, 0, '', sz,  4, "",   "", False, False, 'center', ""]
            , 'font':  [0, 0, '', sz, 20, "",   "", False, False, 'left',   ""]
            , 'sz':    [0, 0, '', sz,  4, "",   "", False, False, 'center', ""]
            , 'w':     [0, 0, '', sz,  4, "",   "", False, False, 'center', ""]
            , 't_clr': [0, 0, '', sz, 10, "",   "", False, False, 'center', ""]
            , 'fill_clr': [0, 0, '', sz, 10, "",   "", False, False, 'center', ""]
            , 'Bold':  [0, 0, '', sz,  7, "",   "", False, False, 'center', ""]
            , 'Ital':  [0, 0, '', sz,  7, "",   "", False, False, 'center', ""]
            , 'Align': [0, 0, '', sz,  7, "",   "", False, False, 'center', ""]
            , 'val':   [0, 0, '', sz, 20, "",   "", False, False, 'left',   ""]
        }

        return a52_def

    # ============================================================================
    def a51_dump_colors(self, cfg, wb, colors_tab):
        # Dump colors
        # These are the colors I used in v_chk 0.5
        COLOR_INDEX = {
            'pur44': [self.Colors.clr_pur44, self.Colors.clr_blk00],
            'pur45': [self.Colors.clr_pur45, self.Colors.clr_wht00],
            'ora64': [self.Colors.clr_ora64, self.Colors.clr_blk00],
            'blud4': [self.Colors.clr_blud4, self.Colors.clr_blk00],
            'red20': [self.Colors.clr_red20, self.Colors.clr_wht00],
            'grn30': [self.Colors.clr_grn30, self.Colors.clr_blk00],
            'aqu56': [self.Colors.clr_aqu56, self.Colors.clr_blk00],
            'gry35': [self.Colors.clr_gry35, self.Colors.clr_blk00],
            'wht15': [self.Colors.clr_wht15, self.Colors.clr_blk00],
            'wht00': [self.Colors.clr_wht00, self.Colors.clr_blk00],
            'blk00': [self.Colors.clr_blk00, self.Colors.clr_wht00],
        }

        tables_start_row = 7
        tables_start_col = 4
        tables_gutter = 1
        tables_width = 6

        com_col = tables_start_col
        com_row = 3

        comments = [ 'Hex Vales are listed with the fill color set to that value. The Text'
               , 'will appear with the complement of that color.'
               , 'Warning: If the text starts displaying as black instead of the fill'
               , 'you are running out of memory. Close other spreadsheets, or applications'
               , 'Warning: If the text appears black, you are running'
               ]

        # Build a colors table
        skip = 3
        cmax = 17
        tables_width = 6
        s = 0
        colors = []  # * (cmax + 1)
        for r in range(0, cmax, skip):
            red = r
            if red > 15:
                red = 15
            rx = f"{red:01x}{red:01x}"
            for g in range(0, cmax, skip):
                grn = g
                if grn > 15:
                    grn = 15
                gx = f"{grn:01x}{grn:01x}"
                for b in range(0, cmax, skip):
                    blu = b
                    if blu > 15:
                        blu = 15
                    bx = f"{blu:01x}{blu:01x}"
                    # step_cnt += 1
                    # if step_cnt % step == 0:
                        # if s % skip == 0:
                    colors.append(f"{rx}{gx}{bx}")
                    print(f"i:{r}  j:{g}  k:{b}   {rx}{gx}{bx}")
                    s += 1

        # Export Pretty Color Table
        row_idx = tables_start_row
        col_idx = tables_start_col
        start_row = row_idx
        start_col = col_idx

        hdr_top_done = False
        hdr_lft_done = False

        cell = colors_tab.cell(row=start_row - 4, column=start_col - 3, value="skip:")
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = colors_tab.cell(row=start_row - 3, column=start_col - 3, value="cmax")
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = colors_tab.cell(row=start_row - 4, column=start_col - 2, value=skip)
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = colors_tab.cell(row=start_row - 3, column=start_col - 2, value=cmax)
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = colors_tab.cell(row=start_row - 1, column=start_col - 2, value="Colors:")
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell = colors_tab.cell(row=start_row - 1, column=start_col - 1, value=len(colors))
        cell.font = Font(bold=True, size=12, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")


        # Export Comments
        for com_row, comment in enumerate(comments, start=com_row):
            cell = colors_tab.cell(row=com_row, column=com_col, value=comment)
            cell.font = Font(size=10, color="000000")

        max_col = 6
        max_row = 36
        r_skip = 6

        # Export Pretty Color Table
        for clr_set in range(0, r_skip):
            for c_row in range(0, max_row, r_skip):
                print(f"clr_set: {clr_set}  row: {c_row + clr_set:2d}: ", end="")

                for c_col in range(0, max_col):
                    if c_row < max_row:
                        # =($K5*5)+L$4+$K5  ($K5 = c_row,   L$4 = c_col)  *5=*(max_col - 1)
                        # = (c_row * (max_col - 1)) + (c_row):4d
                        # works, but set always 0: print(f"{(c_row * (max_col - 1)) + (c_col + c_row):4d}   ", end="")
                        print(f"{(c_row * (max_col - 1)) + (c_col + c_row) + (clr_set * r_skip):4d}   ", end="")
                        clr = colors[(c_row * (max_col - 1)) + (c_col + c_row) + (clr_set * r_skip)]
                        txt_clr = complement(clr)
                        cell = colors_tab.cell(row=row_idx, column=col_idx, value=clr)
                        cell.font = Font(bold=True, size=12, color=txt_clr)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = PatternFill(start_color=clr, end_color=clr,
                                                fill_type="solid", fgColor=txt_clr)

                        if not hdr_top_done:
                            hdr_val = col_idx - start_col + 1
                            cell = colors_tab.cell(row=row_idx - 1, column=col_idx, value=hdr_val)
                            cell.font = Font(bold=True, size=12, color="000000")
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        if not hdr_lft_done:
                            lft_val = row_idx - start_row + 1
                            cell = colors_tab.cell(row=row_idx, column=start_col - 1, value=lft_val)
                            cell.font = Font(bold=True, size=12, color="000000")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            hdr_lft_done = True

                        col_idx += 1
                        if col_idx > tables_width + start_col - 1:
                            hdr_top_done = True
                            hdr_lft_done = False
                            row_idx += 1
                            col_idx = start_col
                print("")

        # Export Ordered Color Table
        start_row = tables_start_row
        start_col = tables_start_col + tables_width + tables_gutter + 1
        row_idx = start_row
        col_idx = start_col

        hdr_top_done = False
        hdr_lft_done = False

        for clr in colors:
            txt_clr = complement(clr)
            cell = colors_tab.cell(row=row_idx, column=col_idx, value=clr)
            cell.font = Font(bold=True, size=12, color=txt_clr)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=clr, end_color=clr,
                                     fill_type="solid", fgColor=txt_clr)

            if not hdr_top_done:
                hdr_val = col_idx - start_col + 1
                cell = colors_tab.cell(row=row_idx - 1, column=col_idx, value=hdr_val)
                cell.font = Font(bold=True, size=12, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if not hdr_lft_done:
                lft_val = row_idx - start_row + 1
                cell = colors_tab.cell(row=row_idx, column=start_col - 1, value=lft_val)
                cell.font = Font(bold=True, size=12, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                hdr_lft_done = True

            # cell.font = Font(bold=True, color=color_value[1])
            col_idx += 1
            if col_idx > tables_width + start_col - 1:
                hdr_top_done = True
                hdr_lft_done = False
                row_idx += 1
                col_idx = start_col

    def export_summary(self, cfg, wb, summary_tab, tab_def):
        # ============================================================================
        # Step 1: "Summary" sheet
        # ============================================================================
        tab_id = tab_def['tab_id']
        tab_name = tab_def['tab_name']
        tab_table_style = tab_def['tab_table_style']
        tab_comments = tab_def['tab_comments']
        tab_color = tab_def['tab_color']
        tbl_hdr_row = tab_def['tbl_hdr_row']
        tbl_beg_col = tab_def['tbl_beg_col']
        tbl_end_col = tab_def['tbl_end_col']
        tab_table = tab_def['tab_table']
        tab_table_links_cols = tab_def['tab_table_links_cols']
        summary_tab.sheet_view.showGridLines = False
        cell = summary_tab.cell(row=4, column=3, value="Created")
        cell.font = Font(name='Berlin Sans FB', size=11, bold=True)
        cell = summary_tab.cell(row=5, column=3, value="Log File")
        cell.font = Font(name='Berlin Sans FB', size=11, bold=True)
        cell = summary_tab.cell(row=6, column=3, value="Log No.")
        cell.font = Font(name='Berlin Sans FB', size=11, bold=True)
#       cell = summary_tab.cell(row=4, column=5, value=f"{cfg.c_date}")
        cell.font = Font(size=11, bold=True)
        cell = summary_tab.cell(row=5, column=5, value=f"{cfg.v_chk_cfg_pname}")
        cell.font = Font(size=11, bold=True)
        cell = summary_tab.cell(row=6, column=5, value=f"{cfg.bat_num:04d}")
        cell.font = Font(size=11, bold=True)
#       xl_set_border(summary_tab, "C4:H6", "thin", tab_color)
        xl_set_border(summary_tab, "C10:F12", "thin", tab_color)
#       # get totals
        # pros = {'tags': {'tag1': [file1, file2, file3...],
        #                     'aa': [file1, file2, file3...]},
        #          'links' : {'[[xxx]]':[file1, file2, file3...],
        #                     '[[zzz]]': [file1, file2, file3...]},
        #          'status': {'pval1': [file1, file2, file3...],
        #                     'pval2': [file1, file2, file3...]}}
        # prop_tab_dict = {'links': ['[[xxx]]', '[[yyy]]', ...]
        #              'status': ['pval1', 'pval2', ...] }
        #  tag_tab_list  = ['tag1', 'aa'...]
        #
        # look into   # sum(len(v) for v in pros['tags'].values())
        self.prop_tab_dict = {}  # pname: [list of all values]
        self.tags_tab_list = []
        len_pros = len(cfg.pros)
        # len_tags = len(cfg.pros["tags"])
        len_tags = 0
        tot_pros = 0
        tot_prop_vals = 0
        tot_tags = 0
        # print(f"cfg_pros: {cfg.pros}")
        for prop_name, values_dict in cfg.pros.items():
            if prop_name != "tags":
                tot_pros += 1
            for value_item, value_files_list in values_dict.items():
                value_item_count = len(values_dict)
                prop_vals_list = []
                if prop_name == "tags":
                    self.tags_tab_list.append(value_item)
                    tot_tags += value_item_count
                    continue
                else:
                    if prop_name in self.prop_tab_dict.keys():
                        prop_vals_list = self.prop_tab_dict[prop_name]
#                   prop_vals_list.append(value_item)
#                   self.prop_tab_dict[prop_name] = prop_vals_list
                    tot_prop_vals += value_item_count
#       # Summary headers Quick Analysis, left column Titles
        s_row = 9
        s_col = 4  # Starting in column D (numerical index = 4)
        rows_sort = ["Quick Analysis", "Totals", "Properties", "Tags"]
        for row_idx, row_title in enumerate(rows_sort, start=s_row):
            cell = summary_tab.cell(row=row_idx, column=s_col, value=row_title)
            if row_idx == s_row:
                cell.font = Font(bold=True, size=11, color=self.Colors.clr_blk00)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = Font(bold=True, color=tab_def.clr_wht00)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = PatternFill(start_color=tab_def.clr_pur45, end_color=tab_def.clr_pur45,
                                        fill_type="solid")
        summary_tab.column_dimensions['D'].width = 20
#       # Summary headers Quick Analysis, top row Titles
        s_row = 10
        s_col = 5  # Starting in column D (numerical index = 4)
        cols_sort = ["Count", "Values", "Files"]
        for col_idx, col_title in enumerate(cols_sort, start=s_col):
            cell = summary_tab.cell(row=s_row, column=col_idx, value=col_title)
            cell.font = Font(bold=True, color=tab_def.clr_wht00)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=tab_def.clr_pur45, end_color=tab_def.clr_pur45,
                                    fill_type="solid")
#       # print summary totals values
        cell = summary_tab.cell(row=11, column=5, value=int(tot_pros))
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal="center", vertical="center")
#       cell = summary_tab.cell(row=12, column=5, value=int(tot_tags))
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal="center", vertical="center")
#       cell = summary_tab.cell(row=11, column=6, value=int(tot_prop_vals))
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal="center", vertical="center")

        return


if __name__ == "__main__":
    # vault_path = "E:\o2"  # Change this to your vault path
    # output_file = "obsidian_metadata.xlsx"

    cfg = Config()
    cfg = cfg.read_config(cfg)

    # wb_cfg = Wb_Cfg()

    if cfg:
        print(f"v_chk_xl: Using last saved config: {cfg.v_chk_xls_pname}")
        exporter = ExcelExporter(cfg)
        exporter.export(cfg)

        print(f"v_chk_xl:Loading Spreadsheet: {cfg.wb_exec_path} - {cfg.v_chk_xls_pname}")
        time.sleep(5)

        # pid = Popen([cfg.wb_exec_path, cfg.v_chk_xls_pname]).pid
    else:
        print(f"v_chk_xl: Error reading config in main: {cfg.v_chk_xls_pname}")
        print(f"v_chk_xl: Exiting...")
