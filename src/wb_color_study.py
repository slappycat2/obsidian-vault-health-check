
import sys
import time
from subprocess import Popen

import openpyxl

from openpyxl.styles import Alignment, Font, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from v_chk_cfg_data import *        # s/b color_study_data import *
from v_chk_wb_tabs import *
from v_chk_class_lib import *

def a51_dump_colors(self, cfg, wb, colors_tab):
    # Dump colors
    # These are the colors I used in v_chk 0.5
    COLOR_INDEX = {
        'pur44': [self.Colors.clr_pur44, self.Colors.clr_blk00],
        'pur45': [self.Colors.clr_pur45, self.Colors.clr_wht00],
        'ora64': [self.Colors.clr_ora64, self.Colors.clr_blk00],
        'blud4': [self.Colors.clr_blud4, self.Colors.clr_blk00],
        'red20': [self.Colors.clr_red10, self.Colors.clr_wht00],
        'grn30': [self.Colors.clr_grn30, self.Colors.clr_blk00],
        'aqu56': [self.Colors.clr_aqu56, self.Colors.clr_blk00],
        'gry35': [self.Colors.clr_gry35, self.Colors.clr_blk00],
        'wht15': [self.Colors.clr_wht15, self.Colors.clr_blk00],
        'wht00': [self.Colors.clr_wht00, self.Colors.clr_blk00],
        'blk00': [self.Colors.clr_blk00, self.Colors.clr_wht00],
    }

    tables_start_row = 7
    tables_start_col = 4
    tables_gutter = 1
    tables_width = 6

    com_col = tables_start_col
    com_row = 3

    comments = ['Hex Vales are listed with the fill color set to that value. The Text'
        , 'will appear with the complement of that color.'
        , 'Warning: If the text starts displaying as black instead of the fill'
        , 'you are running out of memory. Close other spreadsheets, or applications'
        , 'Warning: If the text appears black, you are running'
                ]

    # Build a colors table
    skip = 3
    cmax = 17
    tables_width = 6
    s = 0
    colors = []  # * (cmax + 1)
    for r in range(0, cmax, skip):
        red = r
        if red > 15:
            red = 15
        rx = f"{red:01x}{red:01x}"
        for g in range(0, cmax, skip):
            grn = g
            if grn > 15:
                grn = 15
            gx = f"{grn:01x}{grn:01x}"
            for b in range(0, cmax, skip):
                blu = b
                if blu > 15:
                    blu = 15
                bx = f"{blu:01x}{blu:01x}"
                # step_cnt += 1
                # if step_cnt % step == 0:
                # if s % skip == 0:
                colors.append(f"{rx}{gx}{bx}")
                print(f"i:{r}  j:{g}  k:{b}   {rx}{gx}{bx}")
                s += 1

    # Export Pretty Color Table
    row_idx = tables_start_row
    col_idx = tables_start_col
    start_row = row_idx
    start_col = col_idx

    hdr_top_done = False
    hdr_lft_done = False

    cell = colors_tab.cell(row=start_row - 4, column=start_col - 3, value="skip:")
    cell.font = Font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell = colors_tab.cell(row=start_row - 3, column=start_col - 3, value="cmax")
    cell.font = Font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = colors_tab.cell(row=start_row - 4, column=start_col - 2, value=skip)
    cell.font = Font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell = colors_tab.cell(row=start_row - 3, column=start_col - 2, value=cmax)
    cell.font = Font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    cell = colors_tab.cell(row=start_row - 1, column=start_col - 2, value="Colors:")
    cell.font = Font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell = colors_tab.cell(row=start_row - 1, column=start_col - 1, value=len(colors))
    cell.font = Font(bold=True, size=12, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Export Comments
    for com_row, comment in enumerate(comments, start=com_row):
        cell = colors_tab.cell(row=com_row, column=com_col, value=comment)
        cell.font = Font(size=10, color="000000")

    max_col = 6
    max_row = 36
    r_skip = 6

    # Export Pretty Color Table
    for clr_set in range(0, r_skip):
        for c_row in range(0, max_row, r_skip):
            print(f"clr_set: {clr_set}  row: {c_row + clr_set:2d}: ", end="")

            for c_col in range(0, max_col):
                if c_row < max_row:
                    # =($K5*5)+L$4+$K5  ($K5 = c_row,   L$4 = c_col)  *5=*(max_col - 1)
                    # = (c_row * (max_col - 1)) + (c_row):4d
                    # works, but set always 0: print(f"{(c_row * (max_col - 1)) + (c_col + c_row):4d}   ", end="")
                    print(f"{(c_row * (max_col - 1)) + (c_col + c_row) + (clr_set * r_skip):4d}   ", end="")
                    clr = colors[(c_row * (max_col - 1)) + (c_col + c_row) + (clr_set * r_skip)]
                    txt_clr = Colors.complement(clr)
                    cell = colors_tab.cell(row=row_idx, column=col_idx, value=clr)
                    cell.font = Font(bold=True, size=12, color=txt_clr)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color=clr, end_color=clr,
                                            fill_type="solid", fgColor=txt_clr)

                    if not hdr_top_done:
                        hdr_val = col_idx - start_col + 1
                        cell = colors_tab.cell(row=row_idx - 1, column=col_idx, value=hdr_val)
                        cell.font = Font(bold=True, size=12, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    if not hdr_lft_done:
                        lft_val = row_idx - start_row + 1
                        cell = colors_tab.cell(row=row_idx, column=start_col - 1, value=lft_val)
                        cell.font = Font(bold=True, size=12, color="000000")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        hdr_lft_done = True

                    col_idx += 1
                    if col_idx > tables_width + start_col - 1:
                        hdr_top_done = True
                        hdr_lft_done = False
                        row_idx += 1
                        col_idx = start_col
            print("")

    # Export Ordered Color Table
    start_row = tables_start_row
    start_col = tables_start_col + tables_width + tables_gutter + 1
    row_idx = start_row
    col_idx = start_col

    hdr_top_done = False
    hdr_lft_done = False

    for clr in colors:
        txt_clr = Colors.complement(clr)
        cell = colors_tab.cell(row=row_idx, column=col_idx, value=clr)
        cell.font = Font(bold=True, size=12, color=txt_clr)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=clr, end_color=clr,
                                fill_type="solid", fgColor=txt_clr)

        if not hdr_top_done:
            hdr_val = col_idx - start_col + 1
            cell = colors_tab.cell(row=row_idx - 1, column=col_idx, value=hdr_val)
            cell.font = Font(bold=True, size=12, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if not hdr_lft_done:
            lft_val = row_idx - start_row + 1
            cell = colors_tab.cell(row=row_idx, column=start_col - 1, value=lft_val)
            cell.font = Font(bold=True, size=12, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            hdr_lft_done = True

        # cell.font = Font(bold=True, color=color_value[1])
        col_idx += 1
        if col_idx > tables_width + start_col - 1:
            hdr_top_done = True
            hdr_lft_done = False
            row_idx += 1
            col_idx = start_col

if __name__ == "__main__":
    # vault_path = "E:\o2"  # Change this to your vault path
    # output_file = "obsidian_metadata.xlsx"

    cfg = Config()
    cfg = cfg.read_config(cfg)

    # wb_cfg = Wb_Cfg()

    if cfg:
        print(f"v_chk_xl: Using last saved config: {cfg.v_chk_xls_pname}")
        exporter = ExcelExporter(cfg)
        exporter.export_wb(cfg)

        print(f"v_chk_xl:Loading Spreadsheet: {cfg.xl_exec_path} - {cfg.v_chk_xls_pname}")
        time.sleep(5)

        # pid = Popen([cfg.xl_exec_path, cfg.v_chk_xls_pname]).pid
    else:
        print(f"v_chk_xl: Error reading config in main: {cfg.v_chk_xls_pname}")
        print(f"v_chk_xl: Exiting...")
