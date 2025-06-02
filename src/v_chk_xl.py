import sys
import time
from subprocess import Popen
import urllib.parse
import copy
import re

import openpyxl


from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet import cell_range
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from v_chk_wb_setup import *
# from v_chk_xl_tabs import *
from v_chk_class_lib import *


# WIP
# Todo: Bug-019 - File Count in Properties is really, meaningless. it's files * values
#                 Look for more like this!

# Open
# Todo: Bug-006 - Testing: Mac, Sheets, LibreOffice
# Todo: Bug-014 - make sure all Obsidian and Dataview Prop/Tag rules apply
# Todo: Bug-016 - fix how IsVisCol column is defined and calc'd (it's in 3 places!)
# Todo: Bug-018 - Needs better font support
# Todo: Bug-020 - Exclude templates and Nests from Vault Tabs (props, tags, files, etc.)
# Todo: Bug-022 - Remove this section and track bugs and ERs in Github
# Todo: Bug-0 -

# Todo: ER-002 - Make last file links column show "More exist!"
# Todo: ER-003 - Gather more stats?
#   - Most Tags
#   - Inline vs Frontmatter
#   - Top Tags Graph?
#   - Dataview Stats
# Todo: ER-007 - Support for nested-tags (ie, i/article is two tags, but 1 (e)xplict tag)
#      file.tags: A list of all unique tags in the note.
#        Subtags are broken down by each level, so #Tag/1/A will be stored in the list
#        as [#Tag, #Tag/1, #Tag/1/A].
#      file.etags: A list of all explicit tags in the note;
#        unlike file.tags, does not break subtags down, i.e. [#Tag/1/A] <- so this w/b 1 tag
# Todo: ER-008 - Handle sub-tags (eg, assets/mac/software) better
# Todo: ER-009 - Task shorthands are not supported
#      (see https://blacksmithgu.github.io/obsidian-dataview/annotation/metadata-tasks/#field-shorthands)
# Todo: ER-011 - Array Formulas in Summary? Can't think of one, now, but...
# Todo: ER-012 - Include a Flag to Display Relative Path or just NoteName in all Hyperlinks
# Todo: ER-013 - Identify Singular and Plural usages of properties and tags
# Todo: ER-014 - Fix Area51 Table Dump
# Todo: ER-015 - Rename dirs_skip_rel_str to dirs_skip_abs_lst_user
# Todo: ER-0 -
# Todo: ER-0 -
# Todo: ER-0 -
# Todo: ER-0 -

# Todo: ER-999 - Refactoring:
#   - Use Class sub-classes for tab definitions? Where is there overlap?
#   - Clean up code!!!

# Done
# Todo: Bug-023 - Highlight use of uppercase. Done. (Can only be done in Files)
# Todo: ER-001 - Write GUI for Config
# Todo: ER-010 - Stats: Deprecated Props (alias, cssclass, etc)-Noted in red/italic only, not totals
# Todo: Bug-017 - Move Properties Summary to a new tab
# Todo: Bug-012 - Unique Values calc as diff on Tags when all are showing
# Todo: Bug-013 - make sure inline bookmarks are not seen as tags
# Todo: Bug-009 - Change export_cell to accept relative row, coll
# Todo: Bug-002 - Fix left column title descriptions in Summary
# Todo: Bug-004 - Disable Output, if DBUG_LVL = 0
# Todo: Bug-007 - xkeys appear in all tabs; xkeys is gone
# Todo: Bug-010 - Change "xkey" to something like, x-k-e-y, so it's less likely to dup a prop?
# Todo: Bug-001 - Cleanup/offset tables in Dups and Xyml
# Todo: Bug-010 - Isvisible formula broken in Summary (Check others!)
# Todo: Bug-011 - All tags m/b lowercase
# Todo: Bug-021 - The last row in pros does not show Year, like pros does
#                   Last row may be missing in others (can't reproduce anymore)
# Todo: Bug-011 - Summary Table Totals are wrong
# Todo: Bug-012 - Summary Table contains Tags
# Todo: ER-004 - Research ways to sell it, or just go with Ko-Fi: Ko-Fi
# Todo: ER-006 - List Mapwithtags (Plugins) Property's Separate Tab?
# Todo: Bug-005 - De-couple Area51
# Todo: Bug-015 - skip templates folder (duplicate of 020)
# Todo: Bug-003 - Fix Bad Colors in tbl hdrs (Pending Plugins)
# Todo: Bug-008 - Fix table color in xyml tab


# file.tags: A list of all unique tags in the note.
#   Subtags are broken down by each level, so #Tag/1/A will be stored in the list as [#Tag, #Tag/1, #Tag/1/A].
# file.etags: A list of all explicit tags in the note;
#   unlike file.tags, does not break subtags down, i.e. [#Tag/1/A] <- so this w/b 1 tag


# This is defined outside of the Class, to make it easier to use in other scripts.
# Need to restructure this. Should this be a "library", or maybe part of a WB_CLASS?

class ExcelExporter:
    def __init__(self, dbug_lvl):
        self.DBUG_LVL = dbug_lvl
        self.tab_def = {}
        self.dbug = False
        self.wb_tabs_open = {}
        self.c_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.TAG_TOT_COLS = 20  # Set this to the maximum columns to be added to table (w/o the step)
        self.TAG_BEG_COL = 7  # set the to the first column, ie, where to start
        self.PROP_TOT_COLS = 13  # Set this to the maximum columns to be added to table (w/o the step)
        self.PROP_BEG_COL = 9  # set the to the first column, ie, where to start
        self.COL_STEP = 2  # This should be the number of cols that will be printed in each loop
        self.TABLE_GROUPINGS = False
        self.OPEN_ON_CREATE = True
        self.next_cell_col = 0
        self.last_cell_row = 0
        self.colors = Colors()
        self.DBUG_TAB = 'summ'  # DBUG_LVL must be greater than 2
        # self.DBUG_LVL = 0  # Do Not print anything
        # self.DBUG_LVL = 1  # print report level actions only (export, load, save, etc.) + all lower levels
        # self.DBUG_LVL = 3  # print export_tab + all lower levels
        # self.DBUG_LVL = 4  # print hdr records + all lower levels
        # self.DBUG_LVL = 5  # print detail records + all lower levels
        # self.DBUG_LVL = 9  # print everything (includes export_cell!)

        self.tab_id_sub_key = ''
        self.tabs_built = {}
        self.wb_tabs_done = {}

        cfg_setup = WbDataDef(self.DBUG_LVL)
        self.xyml_descs = cfg_setup.xyml_descs

        self.wb_def = cfg_setup.read_bat_data()
        self.cfg = self.wb_def.get('cfg', {})

        self.tab_seq = self.cfg['tab_seq']   # ['summ', 'pros', 'tags', 'dups', 'xyml']  # ['summ', 'pros', 'tags', 'dups']
        self.vault_path = self.cfg['vault_path']
        self.vault_id = self.cfg['vault_id']
        self.xls_pname = self.cfg['xls_pname']
        if self.DBUG_LVL >= 0:
            print(f"Building workbook {self.xls_pname}...")
        self.bat_pname = self.cfg['cfg_pname']
        self.bat_num = self.cfg['bat_num']
        self.wb_exec_path = self.cfg['wb_exec_path']
        self.wb_data = self.wb_def.get('wb_data', {})
        self.obs_props = self.wb_data.get('obs_props', {})
        self.obs_atags = self.wb_data.get('obs_atags', {})
        self.obs_xyaml = self.wb_data.get('obs_xyaml', {})
        self.obs_dupfn = self.wb_data.get('obs_dupfn', [])
        self.obs_files = self.wb_data.get('obs_files', [])
        self.obs_tmplt = self.wb_data.get('obs_tmplt', [])
        self.obs_codes = self.wb_data.get('obs_codes', [])
        self.obs_nests = self.wb_data.get('obs_nests', [])
        self.obs_nests = self.wb_data.get('obs_plugs', {})

        self.rgx_boundary = re.compile('^---\\s*$', re.MULTILINE)
        self.rgx_body = re.compile('(^|(\\[))([)([A-Za-z0-9_]+)[:]{2}(.*?)(\\]?\\]?)($|\\])')
        self.rgx_tag_pattern = re.compile('#(\w+)')
        self.rgx_noTZdatePattern = r"([0-9]{4})[-\/]([0-1]?[0-9]{1})[-\/]([0-3])?([0-9]{1})(\s+)([0-9]{2}:[0-9]{2}:[0-9]{2})(.*)"
        self.rgx_noTZdateReplace = r"\1-\2-\3\4 \6"
        self.rgx_sub_strip_code_blocks = r'```[\s\S]*?```'
        self.rgx_sub_strip_inline_code = r'`[^`]*`'

        # self.code_q_types = ['TABLE', 'LIST', 'TASK', 'CALENDAR']
        self.plugin_lib = PluginMan(self.vault_path)

        if self.DBUG_LVL > 8:
            print(f"ExcelExport - cfg.wb_exec_path: {self.wb_exec_path}")

        self.exl_file = Path(self.wb_exec_path)

    def export(self, dbug_lvl):
    # =================================================================================
        self.DBUG_LVL = dbug_lvl
        if self.DBUG_LVL > 1:
            print(f"ExcelExport.export - wb_exec_path: {self.wb_exec_path}")

        # Create the workbook instance
        wb = openpyxl.Workbook()

        wb = self.initialize_all_tabs(wb)

        for tab_id in self.tab_seq:
            if self.DBUG_LVL > 3:
                print(f"Exporting tab: {tab_id}")

            self.tab_def = self.wb_def['wb_tabs'][tab_id]
            # self.tab_def = self.wb_tabs_open[tab_id]
            self.tabs_built[tab_id] = self.tab_def
            if tab_id != 'ar51':
                self.wb_tabs_done[tab_id] = self.export_tab(wb)

        self.export_area51(self, wb)

        if self.DBUG_LVL >= 0:
            print(f"Building workbook completed successfully.")

        self.save_workbook(wb)

    def export_area51(self, wb_obj, wb):
        tab_id = 'ar51'
        tab = self.wb_tabs_open[tab_id]

        tab_def = wb_obj.wb_def['wb_tabs'][tab_id]

        img = Image('../img/Area51.png')
        # add to worksheet and anchor next to cells
        tab.add_image(img, 'A1')

        # ========================================================================
        # export Totals Grid, both headers and formulas for totals
        # ========================================================================
        val = ''
        # row_idx = 19
        tot_table = tab_def['tab_cd_fixed_grid']

        for tot_key, tot_col_def_list in tot_table.items():
            if self.DBUG_LVL > 3: # and self.DBUG_TAB == self.tab_def['tab_id']:
                print(f"Grid-Tab: '{tab_id} Val:{tot_key}' {tot_col_def_list}")
                # ,{row_idx} val:{val}  def_val: {def_val}")  # Set a breakpount on this line

            # def export_cell(self, tab, col_def_list, val, row_idx):
            #     col_idx, col_w, txt_clr, fill_clr, bold_bool, align_val = col_def_list
            #   col, row, sz, t_clr, fill_clr, Bold, Italic, Align, Value          len=9
            #   col, w, t_clr, fill_clr, Bold, Align                               len=6
            row_idx = 0
            row_idx, cell = self.export_cell(tab, tot_col_def_list, val, row_idx)

        # ========================================================================
        # export cfg
        # ========================================================================
        cfg_cd_def = self.tab_def['cfg-dump']
        col_idx = cfg_cd_def[0]
        row_idx = cfg_cd_def[1]
        col_sav = col_idx

        for key, value in self.cfg.items():
            if isinstance(value, (list, tuple)):
                val = ' '.join([str(item) for item in value])
            elif isinstance(value, dict):
                val = str(value)
            else:
                val = value

            _, cell = self.export_cell(tab, cfg_cd_def, key, row_idx)
            cfg_cd_def[0] = 0
            _, cell = self.export_cell(tab, cfg_cd_def, val, row_idx)
            cfg_cd_def[0] = col_sav
            row_idx += 1


    def export_tab(self, wb):
        tab_id = self.tab_def['tab_id']
        tab_def = self.tab_def
        tab = self.wb_tabs_open[tab_id]

        tab_table_style = tab_def['tab_table_style']
        tbl_hdr_row = tab_def['tbl_hdr_row']
        tbl_beg_col = tab_def['tbl_beg_col']
        tbl_end_col = tab_def['tbl_end_col']
        tab_table_links_cols = tab_def['tab_table_links_cols']
        tab_tots_isVisible_col = tab_def['tab_tots_isVisible_col']
        tbl_name = tab_def['tbl_name']
        tab_cd_table_hdr = tab_def['tab_cd_table_hdr']
        tab_cd_table_dtl = tab_def['tab_cd_table_dtl']
        hdr_IsVis        = tab_def['hdr_IsVis']
        showGridLines    = self.tab_def['showGridLines']
        data_src           = tab_def['data_src']

        if self.DBUG_LVL > 2 and self.DBUG_TAB == self.tab_def['tab_id']:
            print(f"Exporting Tab: {tab_id}") #  Set a breakpount on this line

        # Create the tab, or rename Sheet 1, in the Summaries case...
        # if tab_id == 'summ':
        #     tab = wb.active
        #     tab.title = tab_name
        # else:
        #     tab = wb.create_sheet(title=tab_name)  # Excel allows max 31 characters in tab names

        # tab.sheet_properties.tabColor = tab_color
        # tab.sheet_view.showGridLines = showGridLines

        # TAB HEADINGS export the Tabs Main Title in a large font
        temp_def = self.tab_def['tab_cd_title_def']
        row_idx = temp_def[1]
        _, cell = self.export_cell(tab, temp_def, '', row_idx)

        # help_txt, subtitles, notes, etc.
        if 'tab_help_txt' in tab_def:
            for hlp_key, hlp_texts in tab_def['tab_help_txt'].items():
                texts_def = tab_def[f'tab_cd_{hlp_key}_def']
                for row_idx, text_line in enumerate(hlp_texts, start=texts_def[1]):
                    _, cell = self.export_cell(tab, texts_def, text_line, row_idx)

        # TABLE-Hdr export the main body table - column headers
        col_idx = tbl_beg_col
        row_idx = tbl_hdr_row
        for hdr_key, hdr_col_def_list in tab_cd_table_hdr.items():

            # unpack header definitions
            col_idx, cell = self.export_cell(tab, hdr_col_def_list, '', tbl_hdr_row)
            if self.DBUG_LVL > 8:
                print(f"Processing [Header]:{hdr_key}")
                print(f" {tab_id}  row:{tbl_hdr_row}  col: {col_idx}  list: {hdr_col_def_list}")

        if tbl_beg_col != 0:
            tab[f"{self.xl_a_col(tbl_beg_col + 1)}{tbl_hdr_row + 1}"] = "Nothing Found."

        # TABLE-Dtl Begin Detail
        row_idx = tbl_hdr_row + 1
        p_v_Index_p_count, p_v_Index_v_count = 0, 0
        beg_prop_group = row_idx + 1
        last_prop_name = ""
        pros_total = values_total = files_total = 0

        # ==============================================================================
        # TABLE-Dtl: Gathering data--Loop each prop_name, getting the list of all values
        # ==============================================================================
        # First, get the tabs data_src
        tab_data_src = self.wb_def['wb_data']
        for src in data_src:
            tab_data_src = tab_data_src[src]

        for prop_name, values_dict in sorted(tab_data_src.items()):

            # filter props for those belonging to this tab only
            # if tab_id == 'dups' and prop_name != 'xkey_dups':
            #     continue
            # if tab_id == 'xyml' and 'xkey_xyml' not in prop_name:
            #     continue
            # if (tab_id == 'pros' or tab_id == 'vals') and prop_name == "tags":
            #     continue
            # if tab_id == 'tags' and prop_name != "tags":
            #     continue
            if tab_id == 'summ':
                continue

            # so for dups, values_dict will be all pathnames using this filename
            prop_name = self.xl_clean_cell(prop_name)
            p_v_Index_p_count += 1
            pros_total += 1
            p_v_Index_v_count = 0
            if tab_id == "vals" and last_prop_name == "":
                last_prop_name = prop_name

            # sorted_values_dict = {}
            # sorted_values_list = sorted(values_dict.keys())    # this creates a list of keys, not a dict
            # for value_key in sorted_values_list:       # so, now we need to rebuild the dictionary
            #     sorted_values_dict[value_key] = values_dict[value_key]

            # TABLE-Dtl Now, build vals[], by looping each value,  getting a list of Files-Where-Used
            # for value_item, value_files_list in values_dict.items():
            for value_item, value_files_list in values_dict.items():
                if tab_id == 'dups' and len(value_files_list) == 1:
                    continue
                # reset tab_def to default, in case any changes were made on the prev row
                this_row_dtl = copy.deepcopy(tab_cd_table_dtl)
                # First, define what the values are going to be for this row
                value_item_count = len(value_files_list)
                vals = []
                p_v_Index_v_count += 1
                values_total += 1
                files_total += value_item_count

                # TABLE Dtl - Set up First (Fixed) Columns Values
                if tab_id == "pros":
                    vals = [int((row_idx - tbl_hdr_row))
                    , last_prop_name
                    , values_total
                    , files_total
                        ]
                elif tab_id == "vals":
                    vals = [int((row_idx - tbl_hdr_row))
                        , prop_name
                        , value_item
                        , value_item_count
                        , f"{p_v_Index_p_count:03d}-{p_v_Index_v_count:05d}"
                            ]
                elif tab_id == "tags":
                    vals = [int((row_idx - tbl_hdr_row))
                        , value_item
                        , value_item_count
                            ]
                elif tab_id == "dups":
                    vals = [int((row_idx - tbl_hdr_row))
                        , self.obs_hyperlink(Path(value_item).name)
                        , value_item_count
                            ]
                elif tab_id == "xyml":
                    vals = [int((row_idx - tbl_hdr_row))
                            , self.obs_hyperlink(Path(value_item).name)
                            , self.xyml_descs[prop_name][0]
                            ]
                elif tab_id == "file":
                    file_nm, loc = prop_name.split("|")
                    if loc == 'F':
                        loc = ''
                    else:
                        loc = '(Inline)'

                    act_prop = value_files_list[0]
                    if act_prop == value_item:
                        act_prop = ""
                    prop_vals = " | ".join(map(str, value_files_list[1:]))

                    vals = [int((row_idx - tbl_hdr_row))
                            , self.obs_hyperlink(Path(file_nm).name)
                            , ' '
                            , loc
                            , value_item
                            , act_prop
                            , len(value_files_list) - 1
                            , prop_vals
                            ]
                elif tab_id == "code":
                    file_nm = prop_name
                    cb_sig = value_item
                    plugin_id = self.plugin_lib.get_name(cb_sig)

                    vals = [int((row_idx - tbl_hdr_row))
                            , self.obs_hyperlink(Path(file_nm).name)
                            , plugin_id
                            , cb_sig
                            , len(value_files_list)
                            ]
                elif tab_id == 'nest':
                    plugin_id, file_nm = prop_name.split("|")
                    vtot = len(value_files_list)
                    prop_vals = " | ".join(map(str, value_files_list[1:]))
                    vals = [int((row_idx - tbl_hdr_row))
                            , plugin_id
                            , self.obs_hyperlink(Path(file_nm).name)
                            , value_item
                            , vtot
                            , prop_vals
                            ]
                elif tab_id == 'plug':
                    # , 0 'id':                     # - cmdr (not used; overwritten with RowId)
                    # , 1 'name':                   # - Commander
                    # , 2 'enable':                 # - true
                    # , 3 'version':                # - 0.5.2
                    # , 4 'minAppVersion':          # - 1.4.0
                    # , 5 'author':                 # - bob, carol, ted & alice
                    # , 6 'authorUrl':              # - https://github.com/phibr0
                    # , 7 'isDesktopOnly':          # - false
                    # , 8 'Description':            # - Customize your workspace by adding commands
                    # , 9 sig list this plugin uses # []
                    vals = value_files_list
                    vals[0] = int(row_idx - tbl_hdr_row) # overwrite the id, not using it...
                    if "http" in vals[6].lower():
                        vals[6] = self.web_hyperlink(vals[6])
                    if vals[7] is False:
                        vals[7] = ""

                    if vals[2] is False:
                        vals[2] = ""
                        this_row_dtl['name'][8] = True  # set name to red, italic
                        this_row_dtl['name'][5] = self.colors.clr_red

                    if len(vals[9]) > 1:
                        vals[9] = " | ".join(map(str, vals[9]))
                    else:
                        vals[9] = "".join(map(str, vals[9]))

                elif tab_id == "tmpl":
                    vals = [int((row_idx - tbl_hdr_row))
                        , prop_name
                        , value_item
                        , value_item_count
                        , f"{p_v_Index_p_count:03d}-{p_v_Index_v_count:05d}"
                        ]

                # TABLE-Dtl Set up List of Files Used, convert them to obsidian URLs, store in vals[]

                # tab_def["tab_cd_table_links"] =  [9, 0, '', 11, 18, "", "", False, False, 'left'  , '']
                for col_count, file in enumerate(value_files_list, start=1):
                    if col_count > tab_table_links_cols:
                        break

                    obs_link = file
                    if tab_id == "dups":
                        # Need qualified relative path w/o vault path
                        obs_link = self.obs_hyperlink(file.replace(self.vault_path, ""))
                    else:
                        # Just need the MD filename
                        if isinstance(file, str) and file.endswith('.md') and Path(file).is_file():
                            obs_link = self.obs_hyperlink(Path(file).name)

                    vals = vals + [obs_link, " "]

                # TABLE-Dtl Finally, now we have all the col values in vals, so we can add this row...
                if self.DBUG_LVL > 4:
                    print(f"Processing [Property]:{prop_name} - [Value]:{value_item} - [Count]: {value_item_count}")

                if tab_id == "pros" and prop_name == last_prop_name:
                    continue

                # ===========================================================================================
                # TABLE-Sub-Dtl Actual start of a new row print, loop thru tab_table columns and add to sheet
                # ===========================================================================================

                for dummy_key, dtl_col_def_list in this_row_dtl.items():
                    col_idx = dtl_col_def_list[0]
                    if col_idx == 0:
                        col_idx = self.next_cell_col
                    # TABLE-Dtl If there is no value in vals for this col
                    if len(vals) <= (col_idx - tbl_beg_col):
                        continue

                    val = vals[col_idx - tbl_beg_col]

                    if self.DBUG_LVL > 8:
                        print(f"Col: {col_idx}-{dummy_key}\t\t  "
                              f"Val {col_idx - tbl_beg_col}: {vals[col_idx - tbl_beg_col]}")

                    if col_idx == tab_tots_isVisible_col:
                        val = tab_def['f_isVisible']

                    if isinstance(val,  list):
                        val = " | ".join(map(str, val))

                    _, _ = self.export_cell(tab, dtl_col_def_list, val, row_idx)

                    # if col_idx == 5 and prop_name == last_prop_name and prop_name == "pros":
                    #     cell.font = Font(color=self.tab_def.clr_gry35)

                if self.TABLE_GROUPINGS:
                    if prop_name != last_prop_name and last_prop_name != "":
                        if beg_prop_group != row_idx - 1:
                            tab.row_dimensions.group(beg_prop_group, row_idx - 1, hidden=False)
                            # wgh = openpyxl.worksheet.dimensions.DimensionHolder(tab, reference='index', default_factory=None)
                            # wgh.group(beg_prop_group, row_idx - 1, outline_level=1, hidden=False)
                            beg_prop_group = row_idx + 1

                last_prop_name = prop_name
                values_total = 0
                files_total = 0
                row_idx += 1

        # Group last group
        if self.TABLE_GROUPINGS:
            if beg_prop_group != row_idx - 1:
                tab.row_dimensions.group(beg_prop_group, row_idx - 1, hidden=False)

        # tab.row_dimensions.co(outlineLevel=1, worksheet=tab, collapsed=True)
        # create a table
        # Calculate table size
        # last_row = tab.max_row
        # if last_row != tbl_hdr_row + 1:
        #     print(f"last_row: {last_row} <> row_idx - tbl_hdr_row: {row_idx - tbl_hdr_row}")
        #     _ = input(f"Press  Return to continue...")
        if self.tab_def['tbl_beg_col']:
            ret = self.format_as_table(tab, tbl_name, tab_table_style, row_idx)

        last_row = row_idx
        # ===========================================================================================
        # FIXED CELLS-These are non-table items that w/"fixed" cell positions on the worksheet (grid)
        # ===========================================================================================

        vis_key = 'isVisible'
        tot_table = tab_def['tab_cd_fixed_grid']

        if self.tab_def['tab_has_isVisible_col'] and tab_tots_isVisible_col:
            # STEP 1 - Set IsVisible formula in last table column
            # Set isVisible Column Header

            row_num = tbl_hdr_row
            cell_def = tab_def['tab_cd_fixed_grid'][vis_key]

            col, row, font, sz, w, t_clr, fill_clr, Bold, Ital, Align, val = cell_def
            _, cell = self.export_cell(tab, cell_def, hdr_IsVis, row_num)
            row_num += 1

            # Now, export IsVisible Formula to all cells in the IsVisible column
            cell_def[5] = cell_def[6] = ''   # shut off hdr colors in dtl
            for row_num in range(row_num, last_row):
                _, cell = self.export_cell(tab, cell_def, val, row_num)

        # Now, print the rest of the grid

        val = None    # get rid of the last val
        row_idx = 0     # use the values defined in the grid

        # ========================================================================
        # export Totals Grid, misc totals with "fixed" cell addresses
        # ========================================================================
        for tot_key, tot_col_def_list in tot_table.items():
            if tot_key == vis_key:
                continue

            if tot_key.startswith("img"):
                #   0   1    2  3   4  5      6        7   8       9     10
                # [col,row,font,sz, w,t_clr,fill_clr,Bold,Ital,  Align,  val ] = 11
                a_img = Image(tot_col_def_list[10])
                a_cell = f"{self.xl_a_col(tot_col_def_list[0])}{tot_col_def_list[1]}"
                # add to worksheet and anchor next to cells
                tab.add_image(a_img, a_cell)
                continue

            if self.DBUG_LVL > 3:  # and self.DBUG_TAB == self.tab_def['tab_id']:
                print(f"Grid-Tab: '{tab_id} Val:{tot_key}' {tot_col_def_list}")
                # ,{row_idx} val:{val}  def_val: {def_val}")  # Set a breakpount on this line

            _, cell = self.export_cell(tab, tot_col_def_list, val, row_idx)

            val = None

        # ========================================================================
        # export Summary Totals (summ), both headers and formulas for totals
        # ========================================================================
        tot_table = tab_def['tab_cd_fixed_summ']
        for tot_key, tot_col_def_list in tot_table.items():
            if self.DBUG_LVL > 3:  # and self.DBUG_TAB == self.tab_def['tab_id']:
                print(f"Grid-summ-Tab: '{tab_id} Val:{tot_key}' {tot_col_def_list}")
                # ,{row_idx} val:{val}  def_val: {def_val}")  # Set a breakpount on this line

            _, cell = self.export_cell(tab, tot_col_def_list, val, row_idx)

            val = None

        # ========================================================================
        # export borders
        # ========================================================================
        if 'borders' in self.tab_def:
            for _, brdr_parms in self.tab_def['borders'].items():
                # 'footer':["C30:J30", "thin", self.colors.clr_blk]
                self.xl_set_border(tab, brdr_parms)

        return tab

    def export_cell(self, tab, col_def_list, val, row_idx):
        # Todo: Change to add the following rules:
        #   If a col_idx = 0 then increment the column by 1, Note that if row=0, row stays the same
        # col,row,font,sz, w,t_clr,f_clr,Bold,Ital,  Align,  val ] = 11
        col_idx, def_row, c_font, c_sz, col_w, txt_clr, fill_clr, bold_bool, ital_bool, align_val, def_val = col_def_list

        if self.DBUG_LVL > 8 and self.DBUG_TAB == self.tab_def['tab_id']:
            print(f"Exporting Cell (Col,Row): {col_def_list[0]},{row_idx} val:{val}  def_val: {def_val}") #  Set a breakpount on this line

        if self.DBUG_LVL > 8 and self.DBUG_TAB == self.tab_def['tab_id']:
            print(f"Exporting Cell (Col,Row): {col_def_list[0]},{row_idx} val:{val}  def_val: {def_val}") #  Set a breakpount on this line

        if val is None or val == "":
            val = def_val

        hyperlink = False
        if isinstance(val, str) and val.upper().startswith("=HYPERLINK"):
            hyperlink = True

        if isinstance(val, str) and val.endswith(":"):
            txt_clr = self.colors.clr_red
            ital_bool = True
            bold_bool = True

        if isinstance(val, str) and val in ['alias', 'cssclass', 'tag']:
            txt_clr = self.colors.clr_red
            ital_bool = True
            bold_bool = True
            val += ' (Deprecated in 1.4)'

        if row_idx == 0:
            if def_row != 0:
                row_idx = def_row
            else:
                row_idx = self.last_cell_row

        if col_idx == 0:
            col_idx = self.next_cell_col
            print(f"    row_idx: {row_idx}  col_idx: {col_idx}  val:{val}")

        if c_sz is None or c_sz == 0:
            c_sz = 10

        if c_font == '' or c_font is None:
            c_font = "default"

        fg_clr = txt_clr
        if fill_clr != "" and fill_clr is not None:
            fg_clr = self.colors.get_txt_clr(fill_clr)

        val = self.xl_clean_cell(val)

        cell = tab.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name=c_font, size=c_sz, bold=bold_bool, italic=ital_bool)
        if align_val == 'wrap':
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        else:
            cell.alignment = Alignment(horizontal=align_val, vertical="center")
        if self.DBUG_LVL > 8 and row_idx < 13:
            print(f"Set Cell: at {row_idx}:{col_idx} to:{val}")

        if txt_clr != "" and txt_clr is not None:
            cell.font = Font(name=c_font, size=c_sz, color=txt_clr, bold=bold_bool, italic=ital_bool)

        if hyperlink:
            if txt_clr == "" or txt_clr is None:
                txt_clr = self.tab_def['tab_link_clr']
            cell.font = Font(color=txt_clr, underline='single', bold=True, italic=ital_bool)

        if fill_clr != "":
            cell.fill = PatternFill(start_color=fill_clr, end_color=fill_clr, fill_type="solid", fgColor=fg_clr)

        # col width
        a_col = self.xl_a_col(col_idx)
        if col_w > 0:
            tab.column_dimensions[a_col].width = col_w

        self.last_cell_row = row_idx
        self.next_cell_col = col_idx + 1

        return col_idx, cell

    def initialize_all_tabs(self, wb):
        live_tab_seq = []
        for tab_id in self.tab_seq:
            tab_def = self.wb_def['wb_tabs'][tab_id]
            data_src = tab_def['data_src'][0]
            if len(self.wb_def['wb_data'][data_src]) == 0:
                continue
            else:
                live_tab_seq.append(tab_id)

            tab_name = tab_def['tab_name']
            tab_color = tab_def['tab_color']
            showGridLines = tab_def['showGridLines']
            # if self.DBUG_LVL > 1 and self.DBUG_TAB == self.tab_def[tab_id]:
            #     print(f"Initializing Tab: {tab_id}")  # Set a breakpount on this line

            # Create the tab, or rename Sheet 1, in the Summaries case...
            if tab_id == 'summ':
                tab = wb.active
                # tab.page_setup.orientation = tab.ORIENTATION_LANDSCAPE
                tab.page_setup.paperSize = tab.PAPERSIZE_A3
                tab.title = tab_name
            else:
                tab = wb.create_sheet(title=tab_name)  # Excel allows max 31 characters in tab names

            self.wb_tabs_open[tab_id] = tab

            tab.sheet_properties.tabColor = tab_color
            tab.sheet_view.showGridLines = showGridLines

        self.tab_seq = live_tab_seq

        return wb

    def save_workbook(self, wb):
        if self.DBUG_LVL >= 0:
            print(f"Saving Spreadsheet: {self.xls_pname}")

        # save and load workbook
        if os.path.isfile(self.xls_pname):
            try_again = True
            w_time = 2 # secs
            retry_max = 60
            retry_count = 0
            while try_again:
                try:
                    retry_count += 1
                    os.remove(self.xls_pname)
                    try_again = False
                except PermissionError:
                    print(f"Attempt {retry_count} to save {self.xls_pname}...")
                    if retry_count < retry_max:
                        print(f"File {self.xls_pname} must be closed. Attempt {retry_count + 1} (of {retry_max} max.) will begin in {w_time} seconds...\n")
                        time.sleep(w_time)
                    else:
                        print(f"File {self.xls_pname} must be closed. Max Retries ({retry_max}) exceeded. {__file__} will now exit.")
                        sys.exit(1)

                except Exception as e:
                    raise Exception(f"Error removing file {self.xls_pname}: {e}")

        wb.save(self.xls_pname)

        if self.OPEN_ON_CREATE:
            self.load_workbook()

    def load_workbook(self):
        pid = Popen([self.wb_exec_path, self.xls_pname]).pid
        return pid

    def obs_hyperlink(self, file):
        file_link = f"{urllib.parse.quote(file, safe=':/')}"
        obs_link_text = file.replace(".md", "")
        obs_link = f'=hyperlink("obsidian://open?vault={self.vault_id}&file={file_link}","{obs_link_text}")'

        return obs_link

    def web_hyperlink(self, file):
        file_link = f"{urllib.parse.quote(file, safe=':/')}"
        web_link_text = file
        web_link = f'=hyperlink("{file_link}","{web_link_text}")'

        return web_link

    def format_as_table(self, tab, tbl_nm, tab_tbl_style, tot_rows):
        tbl_beg_col   = self.tab_def['tbl_beg_col']
        tbl_end_col     = self.tab_def['tbl_end_col']
        tbl_hdr_row = self.tab_def['tbl_hdr_row']
        tbl_rng = f"{self.xl_a_col(tbl_beg_col)}{tbl_hdr_row}:{self.xl_a_col(tbl_end_col)}"
        if tot_rows == int((tot_rows - tbl_hdr_row)):
            tbl_rng = f"{tbl_rng}11"
        else:
            tbl_rng = tbl_rng + str(tot_rows - 1)

        if self.DBUG_LVL > 5:
            print(f"tbl_name: {tbl_nm}  tbl_rng: {tbl_rng}")

        tbl = Table(displayName=tbl_nm, ref=tbl_rng)
        tbl_style = TableStyleInfo(name=tab_tbl_style, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = tbl_style
        tab.add_table(tbl)

        return tab

    def xl_clean_cell(self, cell_value):
        """
        Replace illegal XML characters in the given cell value with '@'
        if the cell value is a string. Handle datetime objects correctly.
        """
        # Check and process datetime objects directly
        if isinstance(cell_value, datetime):
            cell_value = cell_value.replace(tzinfo=None)

        # Process strings for illegal characters
        if isinstance(cell_value, str):
            # Strip TZ from dates in string format if applicable
            if re.compile(self.rgx_noTZdatePattern).search(cell_value):
                cell_value = re.sub(self.rgx_noTZdatePattern, self.rgx_noTZdateReplace, cell_value)
            return ILLEGAL_CHARACTERS_RE.sub("z", cell_value)

        return cell_value

    def xl_a_col(self, col_num):
        col_alpha = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_alpha = chr(65 + remainder) + col_alpha
        return col_alpha

    def xl_set_border(self, ws, b_parms):
        """
        Sets borders for a specified cell range in a worksheet with customizable options
        for border type, color, and sides.

        This function applies borders to cells within a given range in an Excel
        worksheet. Users can define attributes such as the type of border (e.g., thin,
        thick), the color of the border, and the sides of the cell to which the border
        is applied (e.g., top, bottom, left, right, h-sides, v-sides, or all).

        If no values are provided for `border_type`, `color`, or `sides`, default
        values are used. The function allows setting borders to all sides of a cell or
        specific sides such as top, bottom, left, or right, including combinations like
        vertical sides or horizontal sides.

        :param ws: The worksheet object to modify.
        :type ws: openpyxl.worksheet.worksheet.Worksheet
        :param b_parms: A tuple containing settings for the border including the cell
            range, border type, color, and the side(s) for which to apply the border.
            The tuple should contain four elements in the following order:
            1. cell_range (str): The range of cells where borders are to be applied,
               specified in Excel range format (e.g., "A1:C3").
            2. border_type (str): The style of the border (e.g., "thin", "thick").
               Defaults to "thin" if not provided.
            3. color (str): The color of the border in hex color format without "#"
               (e.g., "000000" for black). Defaults to "000000" if not provided.
            4. sides (str): The side(s) of cells to apply the border on. Acceptable
               values are "all", "top", "bottom", "left", "right", "v-sides", or
               "h-sides". Defaults to "all" if not provided.
        :type b_parms: tuple
        :return: None
        :rtype: NoneType
        """

        cell_range, border_type, color, sides = b_parms
        if border_type == None or border_type == "":
            border_type = "thin"
        if color == None or color == "":
            color = "000000"
        if sides == None or sides == "":
            sides = "all"

        sides = sides.lower()
        border = Side(border_style=border_type, color=color)

        if sides == "all":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(top=border, bottom=border)
                    if border_first_col:
                        cell.border = Border(top=border, bottom=border, left=border)
                        border_first_col = False
                    if row[-1] == cell:
                        cell.border = Border(top=border, bottom=border, right=border)
        elif sides == "bottom":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(bottom=border)
        elif sides == "top":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(top=border)
        elif sides == "left":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(left=border)
        elif sides == "right":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(right=border)
        elif sides == "v-sides":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(left=border, right=border)
        elif sides == "h-sides":
            for row in ws[cell_range]:
                border_first_col = True
                for cell in row:
                    cell.border = Border(top=border, bottom=border)

        return

if __name__ == "__main__":
    DBUG_LVL = 1

    # cfg_setup = SysConfig()
    # cfg = cfg_setup.cfg
    exporter = ExcelExporter(DBUG_LVL)
    exporter.export(DBUG_LVL)
    # wb = WbDataDef(DBUG_LVL)
    # wbdef = cfg.read_cfg_data()
    # cfg = wbdef['cfg']
    # xls_pname = cfg['xls_pname']
    # wb_exec_path = cfg['wb_exec_path']

    # wb_cfg = Wb_Cfg()

    # if cfg:
    #     print(f"v_chk_xl: Using last saved config: {xls_pname}")
    #     exporter = ExcelExporter(DBUG_LVL)
    #     exporter.export(DBUG_LVL)
#
    #     print(f"v_chk_xl:Loading Spreadsheet: {wb_exec_path} - {xls_pname}")
    #     time.sleep(5)
#
    #     # pid = Popen([wb_exec_path, xls_pname]).pid
    # else:
    #     print(f"v_chk_xl: Error reading config in main: {xls_pname}")
    #     print(f"v_chk_xl: Exiting...")
